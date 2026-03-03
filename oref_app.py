"""
פיקוד העורף — אוסף התראות + ממשק ייצוא
-----------------------------------------
• Thread אחד שולף GetAlarmsHistory.aspx כל 2 דקות (מחזיר 3000 האחרונות, עדכני להיום)
• Thread שני שולף Alerts.json כל 10 שניות (התראות פעילות ברגע זה)
• Flask מגיש ממשק ווב עברי + כפתור ייצוא Excel
"""

import csv
import hashlib
import sqlite3
import threading
import time
import io
import json
import logging
from datetime import datetime
from collections import Counter, defaultdict

import requests
from flask import Flask, send_file, jsonify, render_template_string, request
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

# ─── Config ───────────────────────────────────────────────────────────────────
DB_PATH           = "alerts.db"
LIVE_INTERVAL     = 10    # seconds — Alerts.json (פעיל עכשיו)
HISTORY_INTERVAL  = 120   # seconds — GetAlarmsHistory (כל 2 דקות, בהתאם ל-cache)
PORT              = 5050
IS_CLOUD          = False  # מוגדר ב-main() לפי env vars

OREF_HEADERS = {
    "User-Agent":       ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                         "AppleWebKit/537.36 (KHTML, like Gecko) "
                         "Chrome/122.0.0.0 Safari/537.36"),
    "Referer":          "https://www.oref.org.il/",
    "X-Requested-With": "XMLHttpRequest",
    "Accept":           "application/json",
}
HISTORY_HEADERS = {
    "Referer":          "https://alerts-history.oref.org.il/12481-he/Pakar.aspx",
    "X-Requested-With": "XMLHttpRequest",
    "Accept":           "application/json, text/javascript, */*; q=0.01",
}

# endpoint עדכני עם נתוני היום — מחדש כל 2 דקות
# endpoint עדכני — 3000 האחרונות, מחדש כל 2 דקות
HISTORY_URL = ("https://alerts-history.oref.org.il/Shared/Ajax/"
               "GetAlarmsHistory.aspx?lang=he&mode=1")
# היסטוריה מלאה (אתמול + מוקדם יותר) — קובץ סטטי שמתעדכן פעם ביום
STATIC_HISTORY_URL = ("https://www.oref.org.il/warningMessages/alert"
                      "/History/AlertsHistory.json")
# התראות פעילות ברגע זה
LIVE_URL    = ("https://www.oref.org.il/warningMessages/alert/Alerts.json")
# ארכיון היסטורי מ-GitHub (עדכון יומי על ידי הקהילה)
CSV_SOURCE_URL = ("https://raw.githubusercontent.com/yuval-harpaz/"
                  "alarms/master/data/alarms.csv")

CAT_NAMES = {
    1:  "ירי רקטות וטילים",
    2:  "חדירת כלי טיס עוין",
    3:  "רעידת אדמה",
    4:  "חשש לצונאמי",
    5:  "אירוע חומרים מסוכנים",
    6:  "התרעה ביטחונית",
    7:  "גל חום",
    8:  "תרגיל",
    13: "ביטול / חזרה לשגרה",
    14: "הנחיות פיקוד העורף",
    15: "חדירת כלי טיס",
    101:"ירי רקטות",
}
CAT_COLORS = {
    1:  "FF4444", 2:  "FF8800", 3:  "AA44FF",
    4:  "4488FF", 5:  "FF44AA", 6:  "FF6600",
    13: "44BB44", 14: "FFCC00", 15: "FF8844",
    101:"FF4444",
}
DEFAULT_COLOR = "AAAAAA"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("oref")

# ─── Database ─────────────────────────────────────────────────────────────────
_db_lock = threading.Lock()

def get_db():
    conn = sqlite3.connect(DB_PATH, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    with get_db() as conn:
        # מגרציה: אם קיים DB ישן ללא עמודת rid — מחק ובנה מחדש
        cols = [r[1] for r in conn.execute("PRAGMA table_info(alerts)").fetchall()]
        if cols and "rid" not in cols:
            log.warning("DB ישן זוהה (ללא rid) — מוחק ובונה מחדש...")
            conn.execute("DROP TABLE IF EXISTS alerts")
            conn.commit()

        conn.execute("""
            CREATE TABLE IF NOT EXISTS alerts (
                rid       INTEGER PRIMARY KEY,
                alert_dt  TEXT    NOT NULL,
                city      TEXT    NOT NULL,
                title     TEXT    NOT NULL,
                category  INTEGER NOT NULL,
                cat_desc  TEXT    DEFAULT '',
                source    TEXT    DEFAULT 'history'
            )
        """)
        conn.execute("CREATE INDEX IF NOT EXISTS idx_dt ON alerts(alert_dt)")

        # מגרציה: הוסף עמודת origin אם לא קיימת (נוצלה מייבוא CSV)
        cols = [r[1] for r in conn.execute("PRAGMA table_info(alerts)").fetchall()]
        if "origin" not in cols:
            conn.execute("ALTER TABLE alerts ADD COLUMN origin TEXT DEFAULT ''")
            log.info("מגרציה: עמודת origin נוספה")

        conn.commit()

def _normalize(r, source):
    """Convert API record (either format) to a unified dict."""
    # GetAlarmsHistory format: rid, data, alertDate, category, category_desc
    if "rid" in r:
        return {
            "rid":      r["rid"],
            "alert_dt": r.get("alertDate", ""),
            "city":     r.get("data", ""),
            "title":    r.get("category_desc", ""),
            "category": r.get("category", 0),
            "cat_desc": r.get("category_desc", ""),
            "source":   source,
        }
    # Alerts.json format: alertDate, data, title, category (no rid)
    # Use a synthetic rid based on timestamp+city hash to avoid collisions
    key = f"{r.get('alertDate','')}-{r.get('data','')}-{r.get('title','')}"
    synthetic_rid = int(hashlib.md5(key.encode()).hexdigest()[:8], 16) * -1
    return {
        "rid":      synthetic_rid,
        "alert_dt": r.get("alertDate", ""),
        "city":     r.get("data", ""),
        "title":    r.get("title", ""),
        "category": r.get("category", 0),
        "cat_desc": CAT_NAMES.get(r.get("category", 0), ""),
        "source":   source,
    }

def insert_alerts(rows, source="history"):
    """Insert list of API records; deduplicate by rid."""
    if not rows:
        return 0
    count = 0
    with _db_lock:
        with get_db() as conn:
            for r in rows:
                try:
                    n = _normalize(r, source)
                    conn.execute(
                        "INSERT OR IGNORE INTO alerts "
                        "(rid, alert_dt, city, title, category, cat_desc, source) "
                        "VALUES (?,?,?,?,?,?,?)",
                        (n["rid"], n["alert_dt"], n["city"], n["title"],
                         n["category"], n["cat_desc"], n["source"])
                    )
                    if conn.execute("SELECT changes()").fetchone()[0]:
                        count += 1
                except Exception as e:
                    log.warning("insert error: %s", e)
            conn.commit()
    return count

def _csv_rid(csv_id: str, city: str, time_str: str) -> int:
    """מחשב rid שלילי ייחודי לשורת CSV (זהה לאלגוריתם ב-import_csv.py)."""
    key = f"csv-{csv_id}-{city}-{time_str}"
    return int(hashlib.md5(key.encode()).hexdigest()[:8], 16) * -1


def import_csv_rows(reader, batch_size: int = 1000) -> int:
    """מייבא שורות מ-DictReader של CSV לתוך ה-DB. מחזיר מספר שורות חדשות שנוספו."""
    CSV_CAT_NAMES = {
        0: "", 1: "ירי רקטות וטילים", 2: "חדירת כלי טיס עוין",
        3: "רעידת אדמה", 4: "חשש לצונאמי", 5: "אירוע חומרים מסוכנים",
        6: "התרעה ביטחונית", 7: "גל חום", 8: "תרגיל",
        13: "ביטול / חזרה לשגרה", 14: "הנחיות פיקוד העורף",
        15: "חדירת כלי טיס", 101: "ירי רקטות",
    }

    batch  = []
    total  = 0
    errors = 0

    def flush(b):
        with _db_lock:
            with get_db() as conn:
                conn.executemany(
                    "INSERT OR IGNORE INTO alerts "
                    "(rid, alert_dt, city, title, category, cat_desc, source, origin) "
                    "VALUES (?,?,?,?,?,?,?,?)",
                    b
                )
                conn.commit()

    for row in reader:
        try:
            csv_id   = row.get("id",          "").strip()
            city     = row.get("cities",       "").strip()
            time_str = row.get("time",         "").strip()
            desc     = row.get("description",  "").strip()
            origin   = row.get("origin",       "").strip()
            try:
                category = int(row.get("threat", "0").strip())
            except ValueError:
                category = 0

            rid      = _csv_rid(csv_id, city, time_str)
            cat_desc = CSV_CAT_NAMES.get(category, desc)
            title    = desc if desc else cat_desc

            batch.append((rid, time_str, city, title, category, cat_desc, "csv", origin))
            total += 1

            if len(batch) >= batch_size:
                flush(batch)
                batch = []

        except Exception as e:
            errors += 1
            if errors <= 5:
                log.warning("CSV import row error: %s", e)

    if batch:
        flush(batch)

    if errors:
        log.warning("CSV import: %d שגיאות", errors)
    return total


def query_alerts(where="", params=()):
    with get_db() as conn:
        return conn.execute(
            f"SELECT * FROM alerts {where} ORDER BY alert_dt DESC",
            params
        ).fetchall()

def get_stats(where="", params=()):
    with get_db() as conn:
        total   = conn.execute(
            f"SELECT COUNT(*) FROM alerts {where}", params).fetchone()[0]
        cities  = conn.execute(
            f"SELECT COUNT(DISTINCT city) FROM alerts {where}", params).fetchone()[0]
        newest  = conn.execute(
            f"SELECT MAX(alert_dt) FROM alerts {where}", params).fetchone()[0]
        oldest  = conn.execute(
            f"SELECT MIN(alert_dt) FROM alerts {where}", params).fetchone()[0]
        by_title = conn.execute(
            f"SELECT title, COUNT(*) as n FROM alerts {where} "
            "GROUP BY title ORDER BY n DESC", params
        ).fetchall()
        top_cities = conn.execute(
            f"SELECT city, COUNT(*) as n FROM alerts {where} "
            "GROUP BY city ORDER BY n DESC LIMIT 10", params
        ).fetchall()
    return dict(total=total, cities=cities, newest=newest,
                oldest=oldest, by_title=by_title, top_cities=top_cities)


def _parse_filters(args):
    """Parse URL query params → (WHERE clause, params tuple)."""
    clauses, params = [], []
    date_from = args.get("date_from", "").strip()
    date_to   = args.get("date_to",   "").strip()
    city      = args.get("city",      "").strip()
    types_raw = args.get("types",     "").strip()

    if date_from:
        clauses.append("alert_dt >= ?")
        params.append(date_from + " 00:00:00")
    if date_to:
        clauses.append("alert_dt <= ?")
        params.append(date_to   + " 23:59:59")
    if city:
        clauses.append("city LIKE ?")
        params.append(f"%{city}%")
    if types_raw:
        types = [t.strip() for t in types_raw.split(",") if t.strip()]
        if types:
            placeholders = ",".join("?" * len(types))
            clauses.append(f"title IN ({placeholders})")
            params.extend(types)

    origins_raw = args.get("origins", "").strip()
    if origins_raw:
        origins = [o.strip() for o in origins_raw.split(",") if o.strip()]
        if origins:
            placeholders = ",".join("?" * len(origins))
            clauses.append(f"origin IN ({placeholders})")
            params.extend(origins)

    where = ("WHERE " + " AND ".join(clauses)) if clauses else ""
    return where, tuple(params)

# ─── Collectors ───────────────────────────────────────────────────────────────
_state = {
    "last_live":    "—",
    "last_history": "—",
    "live_errors":  0,
    "new_today":    0,
}

def fetch_json(url, headers=None):
    r = requests.get(url, headers=headers or OREF_HEADERS, timeout=15)
    r.raise_for_status()
    text = r.text.lstrip("\ufeff").strip()
    if not text or text in ("[]", ""):
        return []
    parsed = json.loads(text)
    # וודא שהתוצאה תמיד רשימה
    if isinstance(parsed, list):
        return parsed
    if isinstance(parsed, dict):
        return [parsed]
    return []

def _expand_live(obj):
    """
    Alerts.json מחזיר אובייקט יחיד כשיש התראה פעילה:
    {"id":"1","cat":"1","title":"ירי רקטות","data":["עיר1","עיר2"],...}
    פורש לרשימת רשומות — אחת לכל עיר.
    """
    if not isinstance(obj, dict):
        return []
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    cities  = obj.get("data", [])
    title   = obj.get("title", "")
    try:
        cat = int(obj.get("cat", 0))
    except (ValueError, TypeError):
        cat = 0
    if isinstance(cities, str):
        cities = [cities]
    return [
        {"alertDate": now_str, "data": city, "title": title, "category": cat}
        for city in cities
    ]

def collect_live():
    """Poll Alerts.json every LIVE_INTERVAL seconds (התראות פעילות כרגע)."""
    log.info("Live collector started (every %ds)", LIVE_INTERVAL)
    while True:
        try:
            raw = fetch_json(LIVE_URL, headers=OREF_HEADERS)
            # raw הוא רשימה; כל פריט יכול להיות אובייקט התראה אחת
            records = []
            for item in raw:
                if isinstance(item, dict) and "data" in item and isinstance(item["data"], list):
                    # פורמט Alerts.json — data הוא מערך ערים
                    records.extend(_expand_live(item))
                elif isinstance(item, dict):
                    records.append(item)
                # אחרת (str וכד') — מדלג
            n = insert_alerts(records, source="live")
            _state["last_live"] = datetime.now().strftime("%H:%M:%S")
            if n:
                _state["new_today"] += n
                log.info("Live: +%d new alerts", n)
            _state["live_errors"] = 0
        except Exception as e:
            _state["live_errors"] += 1
            log.warning("Live fetch error: %s", e)
        time.sleep(LIVE_INTERVAL)

def collect_history():
    """Poll GetAlarmsHistory.aspx every HISTORY_INTERVAL seconds.
    מחזיר 3000 ההתראות האחרונות עם rid ייחודי — עדכני להיום."""
    log.info("History collector started (every %ds)", HISTORY_INTERVAL)
    while True:
        try:
            data = fetch_json(HISTORY_URL, headers=HISTORY_HEADERS)
            n = insert_alerts(data, source="history")
            _state["last_history"] = datetime.now().strftime("%H:%M:%S")
            if n:
                _state["new_today"] += n
                log.info("History: +%d new alerts", n)
        except Exception as e:
            log.warning("History fetch error: %s", e)
        time.sleep(HISTORY_INTERVAL)

# ─── Excel Builder ────────────────────────────────────────────────────────────
THIN        = Side(style="thin", color="CCCCCC")
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HDR_FONT    = Font(name="Arial", bold=True, color="FFFFFF", size=11)
HDR_FILL    = PatternFill("solid", fgColor="1F4E79")
CELL_FONT   = Font(name="Arial", size=10)
ALT_FILL    = PatternFill("solid", fgColor="EBF3FB")
RTL         = Alignment(horizontal="right", vertical="center", readingOrder=2)
CTR         = Alignment(horizontal="center", vertical="center", readingOrder=2)

def _hdr(ws, row, cols):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = HDR_FONT; cell.fill = HDR_FILL
        cell.alignment = CTR; cell.border = CELL_BORDER
    ws.row_dimensions[row].height = 22

def _cell(ws, row, col, val, color=None, bold=False, wrap=False):
    c = ws.cell(row=row, column=col, value=val)
    c.font   = Font(name="Arial", size=10, bold=bold,
                    color="FFFFFF" if color else "000000")
    c.border = CELL_BORDER
    c.alignment = Alignment(horizontal="right", vertical="top" if wrap else "center",
                             readingOrder=2, wrap_text=wrap)
    if color:
        c.fill = PatternFill("solid", fgColor=color)
    elif row % 2 == 0:
        c.fill = ALT_FILL
    return c

def build_excel(where="", params=(), filter_info=None):
    rows = query_alerts(where, params)
    wb   = openpyxl.Workbook()

    # ── Sheet 1: All alerts ──────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "כל ההתראות"
    ws1.sheet_view.rightToLeft = True

    for ci, h in enumerate(["חותמת זמן","יישוב","סוג התראה","קטגוריה"], 1):
        ws1.cell(row=1, column=ci, value=h)
    _hdr(ws1, 1, 4)

    for ri, r in enumerate(rows, 2):
        cat   = r["category"]
        color = CAT_COLORS.get(cat, DEFAULT_COLOR)
        _cell(ws1, ri, 1, r["alert_dt"])
        _cell(ws1, ri, 2, r["city"])
        _cell(ws1, ri, 3, r["title"])
        _cell(ws1, ri, 4, CAT_NAMES.get(cat, f"קטגוריה {cat}"),
              color=color, bold=True)

    ws1.auto_filter.ref = f"A1:D{len(rows)+1}"
    ws1.freeze_panes   = "A2"
    for col, w in {1:22, 2:28, 3:38, 4:26}.items():
        ws1.column_dimensions[get_column_letter(col)].width = w

    # ── Sheet 2: Event summary ───────────────────────────────────────────────
    ws2 = wb.create_sheet("סיכום אירועים")
    ws2.sheet_view.rightToLeft = True

    groups  = defaultdict(set)
    g_meta  = {}
    for r in rows:
        ts_min = r["alert_dt"][:16] + ":00"
        key    = (ts_min, r["title"])
        groups[key].add(r["city"])
        g_meta[key] = (r["category"],
                       CAT_COLORS.get(r["category"], DEFAULT_COLOR))

    grp_rows = sorted(
        [(ts, title, sorted(cities), *g_meta[(ts,title)])
         for (ts,title), cities in groups.items()],
        key=lambda x: x[0]
    )

    for ci, h in enumerate(["חותמת זמן","סוג התראה","מספר ישובים","רשימת ישובים"], 1):
        ws2.cell(row=1, column=ci, value=h)
    _hdr(ws2, 1, 4)

    for ri, (ts, title, cities, cat, color) in enumerate(grp_rows, 2):
        _cell(ws2, ri, 1, ts)
        _cell(ws2, ri, 2, title, color=color, bold=True)
        _cell(ws2, ri, 3, len(cities))
        c = ws2.cell(row=ri, column=4, value=", ".join(cities))
        c.font      = CELL_FONT
        c.border    = CELL_BORDER
        c.alignment = Alignment(horizontal="right", vertical="top",
                                 readingOrder=2, wrap_text=True)
        if ri % 2 == 0:
            c.fill = ALT_FILL

    ws2.auto_filter.ref = f"A1:D{len(grp_rows)+1}"
    ws2.freeze_panes    = "A2"
    for col, w in {1:20, 2:38, 3:16, 4:60}.items():
        ws2.column_dimensions[get_column_letter(col)].width = w

    # ── Sheet 3: Statistics ──────────────────────────────────────────────────
    ws3 = wb.create_sheet("סטטיסטיקות")
    ws3.sheet_view.rightToLeft = True

    s   = get_stats(where, params)
    row = 1

    def sect(text):
        nonlocal row
        ws3.merge_cells(start_row=row, start_column=1,
                        end_row=row, end_column=3)
        c = ws3.cell(row=row, column=1, value=text)
        c.font      = Font(name="Arial", bold=True, size=11, color="FFFFFF")
        c.fill      = PatternFill("solid", fgColor="2E75B6")
        c.alignment = Alignment(horizontal="right", vertical="center",
                                 readingOrder=2)
        c.border    = CELL_BORDER
        ws3.row_dimensions[row].height = 20
        row += 1

    def kv(k, v):
        nonlocal row
        ck = ws3.cell(row=row, column=1, value=k)
        cv = ws3.cell(row=row, column=2, value=v)
        for c in (ck, cv):
            c.font      = Font(name="Arial", size=10,
                               bold=(c is ck))
            c.alignment = RTL
            c.border    = CELL_BORDER
        if row % 2 == 0:
            for col in [1,2]:
                ws3.cell(row=row, column=col).fill = ALT_FILL
        row += 1

    ws3.cell(row=row, column=1,
             value="סטטיסטיקות התראות פיקוד העורף").font = \
        Font(name="Arial", bold=True, size=14, color="1F4E79")
    ws3.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    ws3.cell(row=row, column=1).alignment = Alignment(
        horizontal="right", vertical="center", readingOrder=2)
    ws3.row_dimensions[row].height = 30
    row += 2

    sect("סיכום כללי")
    kv('סה"כ התראות', s["total"])
    kv("ישובים מושפעים", s["cities"])
    kv("אירועים מקובצים", len(grp_rows))
    kv("התראה ראשונה", s["oldest"] or "—")
    kv("התראה אחרונה", s["newest"] or "—")
    kv("עודכן לאחרונה",
       datetime.now().strftime("%d/%m/%Y %H:%M:%S"))
    if filter_info:
        kv("פילטר מופעל", filter_info)
    row += 1

    sect("פירוט לפי סוג התראה")
    for ci, h in enumerate(["סוג התראה","מספר","אחוז"], 1):
        c = ws3.cell(row=row, column=ci, value=h)
        c.font = HDR_FONT; c.fill = HDR_FILL
        c.alignment = CTR; c.border = CELL_BORDER
    row += 1
    for ri, r in enumerate(s["by_title"]):
        pct = f"{100*r['n']/s['total']:.1f}%" if s["total"] else "0%"
        for ci, v in enumerate([r["title"], r["n"], pct], 1):
            c = ws3.cell(row=row, column=ci, value=v)
            c.font = CELL_FONT; c.alignment = RTL; c.border = CELL_BORDER
            if ri % 2 == 1:
                c.fill = ALT_FILL
        row += 1

    row += 1
    sect("10 הישובים עם הכי הרבה התראות")
    for ci, h in enumerate(["ישוב","מספר","אחוז"], 1):
        c = ws3.cell(row=row, column=ci, value=h)
        c.font = HDR_FONT; c.fill = HDR_FILL
        c.alignment = CTR; c.border = CELL_BORDER
    row += 1
    for ri, r in enumerate(s["top_cities"]):
        pct = f"{100*r['n']/s['total']:.1f}%" if s["total"] else "0%"
        for ci, v in enumerate([r["city"], r["n"], pct], 1):
            c = ws3.cell(row=row, column=ci, value=v)
            c.font = CELL_FONT; c.alignment = RTL; c.border = CELL_BORDER
            if ri % 2 == 1:
                c.fill = ALT_FILL
        row += 1

    for col, w in {1:38, 2:16, 3:12}.items():
        ws3.column_dimensions[get_column_letter(col)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# ─── Flask App ────────────────────────────────────────────────────────────────
app = Flask(__name__)

HTML = """<!DOCTYPE html>
<html lang="he" dir="rtl">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>פיקוד העורף — ניטור התראות</title>
<style>
  * { box-sizing: border-box; margin: 0; padding: 0; }
  body { font-family: Arial, sans-serif; background: #0d1b2a; color: #e8eaf6; min-height: 100vh; padding: 24px; }
  header { display: flex; align-items: center; gap: 16px; margin-bottom: 28px; border-bottom: 2px solid #1f4e79; padding-bottom: 16px; }
  .logo { width: 54px; height: 54px; background: #1f4e79; border-radius: 50%; display: flex; align-items: center; justify-content: center; font-size: 28px; }
  h1 { font-size: 1.6rem; color: #90caf9; }
  h1 small { font-size: 0.85rem; color: #78909c; font-weight: normal; }

  /* Cards */
  .cards { display: grid; grid-template-columns: repeat(auto-fit, minmax(190px, 1fr)); gap: 14px; margin-bottom: 24px; }
  .card { background: #162032; border: 1px solid #1f4e79; border-radius: 12px; padding: 18px; text-align: center; }
  .card .num { font-size: 2.2rem; font-weight: bold; color: #42a5f5; }
  .card .lbl { font-size: 0.82rem; color: #78909c; margin-top: 4px; }
  .card.green .num { color: #66bb6a; } .card.red .num { color: #ef5350; } .card.yellow .num { color: #ffa726; }

  /* Status bar */
  .status-bar { background: #162032; border: 1px solid #1a3a5c; border-radius: 8px; padding: 10px 18px; margin-bottom: 20px; display: flex; gap: 24px; font-size: 0.83rem; color: #78909c; flex-wrap: wrap; }
  .status-bar span { display: flex; align-items: center; gap: 6px; }
  .dot { width: 8px; height: 8px; border-radius: 50%; background: #66bb6a; animation: pulse 2s infinite; }
  @keyframes pulse { 0%,100%{opacity:1} 50%{opacity:.3} }

  /* Filter panel */
  .filter-panel { background: #162032; border: 1px solid #1f4e79; border-radius: 12px; margin-bottom: 14px; overflow: hidden; }
  .filter-title {
    display: flex; align-items: center; justify-content: space-between;
    padding: 14px 20px; cursor: pointer; user-select: none;
    font-size: .95rem; font-weight: bold; color: #90caf9;
    transition: background .15s;
  }
  .filter-title:hover { background: rgba(255,255,255,.03); }
  .filter-title-right { display: flex; align-items: center; gap: 10px; }
  .filter-active-tag { font-size: .73rem; color: #42a5f5; font-style: italic; font-weight: normal; }
  .filter-toggle { font-size: .8rem; color: #546e7a; transition: transform .2s; }
  .filter-toggle.open { transform: rotate(0deg); }
  .filter-toggle.closed { transform: rotate(-90deg); }

  .filter-body { padding: 0 20px 18px; border-top: 1px solid #1a3a5c; }

  /* Date presets */
  .date-presets { display: flex; gap: 7px; flex-wrap: wrap; margin-bottom: 14px; padding-top: 16px; }
  .preset-btn {
    padding: 5px 13px; background: #0d1b2a; border: 1px solid #1a3a5c;
    border-radius: 20px; color: #90caf9; cursor: pointer;
    font-size: .8rem; font-family: Arial, sans-serif; transition: all .15s;
  }
  .preset-btn:hover { border-color: #42a5f5; color: #42a5f5; background: rgba(66,165,245,.08); }
  .preset-btn.active-preset { border-color: #42a5f5; background: rgba(66,165,245,.15); color: #e8eaf6; }

  /* Filter rows */
  .filter-row { display: flex; gap: 14px; flex-wrap: wrap; margin-bottom: 14px; align-items: flex-end; }
  .filter-group { display: flex; flex-direction: column; gap: 5px; }
  .filter-group label { font-size: .78rem; color: #78909c; }
  .filter-group input[type="date"],
  .filter-group input[type="text"] {
    background: #0d1b2a; border: 1px solid #1a3a5c; border-radius: 6px;
    color: #e8eaf6; padding: 8px 11px; font-size: .88rem; font-family: Arial, sans-serif;
  }
  .filter-group input[type="date"] { min-width: 148px; }
  .filter-group input[type="text"] { min-width: 210px; }
  .filter-group input:focus { outline: none; border-color: #42a5f5; }
  .filter-group input[type="date"]::-webkit-calendar-picker-indicator { filter: invert(.8); }

  /* Type / Origin multi-select */
  #type-select, #origin-select {
    background: #0d1b2a; border: 1px solid #1a3a5c; border-radius: 8px;
    color: #e8eaf6; font-family: Arial, sans-serif; font-size: .85rem;
    width: 100%; padding: 4px; min-height: 130px; resize: vertical;
  }
  #type-select option, #origin-select option { padding: 6px 10px; cursor: pointer; }
  #type-select option:checked, #origin-select option:checked { background: #1565c0; color: #fff; }
  #type-select:focus, #origin-select:focus { outline: none; border-color: #42a5f5; }
  .type-select-hint { font-size: .73rem; color: #546e7a; margin-top: 5px; }

  /* Filter actions */
  .filter-actions { display: flex; gap: 10px; margin-top: 14px; align-items: center; flex-wrap: wrap; }
  .btn-clear { padding: 7px 14px; background: transparent; color: #546e7a; border: 1px solid #37474f; border-radius: 8px; cursor: pointer; font-size: .83rem; font-family: Arial, sans-serif; }
  .btn-clear:hover { color: #e8eaf6; border-color: #78909c; }

  /* Result bar */
  .result-bar { font-size: .82rem; color: #546e7a; margin-bottom: 14px; min-height: 20px; }
  .result-bar.filtered { color: #42a5f5; }
  .result-bar.empty { color: #ef9a9a; }

  /* Export button */
  .export-btn { display: block; width: 100%; padding: 17px; font-size: 1.25rem; font-weight: bold; background: linear-gradient(135deg, #1565c0, #0d47a1); color: #fff; border: none; border-radius: 12px; cursor: pointer; letter-spacing: .5px; transition: all .2s; margin-bottom: 20px; text-decoration: none; text-align: center; }
  .export-btn:hover { background: linear-gradient(135deg, #1976d2, #1565c0); transform: translateY(-2px); box-shadow: 0 6px 20px rgba(21,101,192,.5); }
  .export-btn:active { transform: translateY(0); }

  /* Table */
  table { width: 100%; border-collapse: collapse; font-size: .87rem; margin-top: 18px; }
  thead th { background: #1f4e79; color: #fff; padding: 10px 14px; text-align: right; font-weight: bold; }
  tbody tr:nth-child(even) { background: #162032; } tbody tr:hover { background: #1a3a5c; }
  tbody td { padding: 8px 14px; border-bottom: 1px solid #1a3a5c; }
  .empty-row td { text-align: center; color: #546e7a; padding: 24px; font-size: .88rem; }
  .badge { display: inline-block; padding: 2px 8px; border-radius: 20px; font-size: .77rem; font-weight: bold; color: #fff; }
  .section-title { font-size: 1.05rem; color: #90caf9; margin-bottom: 10px; padding-bottom: 6px; border-bottom: 1px solid #1f4e79; }
  #refresh-notice { font-size: .77rem; color: #546e7a; text-align: center; margin-top: 22px; }
</style>
</head>
<body>

<header>
  <div class="logo">🛡️</div>
  <div>
    <h1>פיקוד העורף — ניטור התראות
      <br><small>נתונים חיים · מתרענן כל 15 שניות</small>
    </h1>
  </div>
</header>

<div class="cards">
  <div class="card red">  <div class="num" id="total">—</div>  <div class="lbl">סה"כ התראות</div></div>
  <div class="card">      <div class="num" id="cities">—</div> <div class="lbl">ישובים מושפעים</div></div>
  <div class="card yellow"><div class="num" id="new_today">—</div><div class="lbl">חדשות מסשן זה</div></div>
  <div class="card green"> <div class="num" id="newest">—</div> <div class="lbl">התראה אחרונה</div></div>
</div>

<div class="status-bar">
  <span><span class="dot"></span> אוסף נתונים חיים</span>
  <span>עדכון (live): <b id="last_live">—</b></span>
  <span>עדכון (היסטוריה): <b id="last_history">—</b></span>
  <span>שגיאות: <b id="errors">0</b></span>
</div>

<!-- ── Filter Panel ── -->
<div class="filter-panel" id="filter-panel">

  <div class="filter-title" onclick="toggleFilterPanel()">
    <span>🔍 סינון נתונים <span class="filter-active-tag" id="filter-active-tag"></span></span>
    <div class="filter-title-right">
      <span id="filter-toggle" class="filter-toggle open">▼</span>
    </div>
  </div>

  <div class="filter-body" id="filter-body">

    <!-- קיצורי תאריך -->
    <div class="date-presets">
      <button class="preset-btn" id="preset-today"     onclick="setPreset('today')">היום</button>
      <button class="preset-btn" id="preset-yesterday" onclick="setPreset('yesterday')">אתמול</button>
      <button class="preset-btn" id="preset-3days"     onclick="setPreset('3days')">3 ימים אחרונים</button>
      <button class="preset-btn" id="preset-week"      onclick="setPreset('week')">שבוע אחרון</button>
      <button class="preset-btn" id="preset-all"       onclick="setPreset('all')">הכל</button>
    </div>

    <div class="filter-row">
      <div class="filter-group">
        <label>מתאריך</label>
        <input type="date" id="f-from">
      </div>
      <div class="filter-group">
        <label>עד תאריך</label>
        <input type="date" id="f-to">
      </div>
      <div class="filter-group" style="flex:1;min-width:210px">
        <label>יישוב</label>
        <input type="text" id="f-city" list="city-list" placeholder="חפש שם יישוב...">
        <datalist id="city-list"></datalist>
      </div>
    </div>

    <div class="filter-row" style="align-items:flex-start">
      <div class="filter-group" style="flex:3;min-width:200px">
        <label>סוגי התראות</label>
        <select multiple id="type-select" size="6">
          <option value="" disabled>טוען...</option>
        </select>
        <div class="type-select-hint">Ctrl+לחיצה לבחירה מרובה &nbsp;·&nbsp; ללא בחירה = כל הסוגים</div>
        <div style="display:flex;gap:6px;margin-top:6px">
          <button class="btn-clear" onclick="selectAllTypes()" style="color:#42a5f5;border-color:#1a3a5c;font-size:.75rem;padding:4px 10px;">✓ הכל</button>
          <button class="btn-clear" onclick="clearTypes()" style="font-size:.75rem;padding:4px 10px;">✕ נקה</button>
        </div>
      </div>
      <div class="filter-group" style="flex:2;min-width:160px">
        <label>מקור <span style="color:#546e7a;font-size:.72rem;font-weight:normal">— מדינה / ארגון מתקיף</span></label>
        <select multiple id="origin-select" size="6">
          <option value="" disabled>טוען...</option>
        </select>
        <div class="type-select-hint">ללא בחירה = כל המקורות</div>
        <div style="display:flex;gap:6px;margin-top:6px">
          <button class="btn-clear" onclick="selectAllOrigins()" style="color:#42a5f5;border-color:#1a3a5c;font-size:.75rem;padding:4px 10px;">✓ הכל</button>
          <button class="btn-clear" onclick="clearOrigins()" style="font-size:.75rem;padding:4px 10px;">✕ נקה</button>
        </div>
      </div>
    </div>

    <div class="filter-actions">
      <button class="btn-clear" onclick="clearFilters()">✕ נקה הכל</button>
    </div>

  </div><!-- /filter-body -->
</div><!-- /filter-panel -->

<!-- Result bar -->
<div class="result-bar" id="result-bar"></div>

<a id="export-btn" href="/export" class="export-btn">⬇️ &nbsp; ייצא קובץ Excel (לפי פילטר)</a>

<div class="section-title">10 התראות אחרונות</div>
<table>
  <thead><tr><th>זמן</th><th>ישוב</th><th>סוג</th></tr></thead>
  <tbody id="recent"></tbody>
</table>

<div id="refresh-notice">מתרענן אוטומטית · <span id="countdown">15</span>ש</div>

<script>
const TYPE_COLORS = {1:"FF4444",2:"FF8800",3:"AA44FF",4:"4488FF",5:"FF44AA",6:"FF6600",13:"44BB44",14:"FFCC00",15:"FF8844",101:"FF4444"};

function esc(s){ return String(s).replace(/&/g,"&amp;").replace(/</g,"&lt;").replace(/>/g,"&gt;").replace(/"/g,"&quot;"); }

function fmtDt(s) {
  // מנרמל גם "YYYY-MM-DD HH:MM:SS" וגם "YYYY-MM-DDTHH:MM:SS"
  const clean = String(s).replace('T', ' ');
  const [date, time] = clean.split(' ');
  const [y, m, d] = (date || '').split('-');
  return `${d}-${m}-${y} ${(time || '').slice(0, 5)}`;
}

function colorFor(t) {
  if (t.includes("רקטות")||t.includes("טילים")) return "#ef5350";
  if (t.includes("ביטול")||t.includes("שגרה")||t.includes("לצאת")) return "#66bb6a";
  if (t.includes("הנחיות")||t.includes("להישאר")||t.includes("סמיכות")) return "#ffa726";
  if (t.includes("כלי טיס")) return "#ff7043";
  if (t.includes("בדקות")) return "#78909c";
  return "#78909c";
}

/* ── Filter state ── */
let allTypes   = [];
let allOrigins = [];
let debounceTimer = null;

const ORIGIN_HE = {
  "Iran":    "איראן",
  "Gaza":    "עזה",
  "Lebanon": "לבנון",
  "Yemen":   "תימן",
  "Iraq":    "עיראק",
  "Syria":   "סוריה",
  "Israel":  "ישראל",
  "FA":      "אזעקת שווא",
};
let filterPanelOpen = true;

/* ── Date presets ── */
const fmt = d => d.toISOString().slice(0,10);
const shiftDay = (d, n) => { const r = new Date(d); r.setDate(r.getDate()+n); return r; };

function setPreset(p) {
  const today = new Date();
  document.querySelectorAll(".preset-btn").forEach(b => b.classList.remove("active-preset"));
  document.getElementById("preset-" + p)?.classList.add("active-preset");
  const from = document.getElementById("f-from");
  const to   = document.getElementById("f-to");
  if      (p === "today")     { from.value = fmt(today);            to.value = fmt(today); }
  else if (p === "yesterday") { const y = shiftDay(today,-1); from.value = fmt(y); to.value = fmt(y); }
  else if (p === "3days")     { from.value = fmt(shiftDay(today,-2)); to.value = fmt(today); }
  else if (p === "week")      { from.value = fmt(shiftDay(today,-6)); to.value = fmt(today); }
  else                        { from.value = ""; to.value = ""; }
  applyFilters();
}

function detectPreset() {
  const from = document.getElementById("f-from").value;
  const to   = document.getElementById("f-to").value;
  const today = fmt(new Date());
  document.querySelectorAll(".preset-btn").forEach(b => b.classList.remove("active-preset"));
  if (!from && !to) document.getElementById("preset-all").classList.add("active-preset");
  else if (from === today && to === today) document.getElementById("preset-today").classList.add("active-preset");
  else if (from === fmt(shiftDay(new Date(),-1)) && to === from) document.getElementById("preset-yesterday").classList.add("active-preset");
  else if (from === fmt(shiftDay(new Date(),-6)) && to === today) document.getElementById("preset-week").classList.add("active-preset");
}

/* ── Collapsible panel ── */
function toggleFilterPanel() {
  filterPanelOpen = !filterPanelOpen;
  document.getElementById("filter-body").style.display = filterPanelOpen ? "" : "none";
  const tog = document.getElementById("filter-toggle");
  tog.classList.toggle("open",   filterPanelOpen);
  tog.classList.toggle("closed", !filterPanelOpen);
}

/* ── Build params ── */
function buildFilterParams() {
  const p = new URLSearchParams();
  const from = document.getElementById("f-from").value;
  const to   = document.getElementById("f-to").value;
  const city = document.getElementById("f-city").value.trim();

  const tSel   = document.getElementById("type-select");
  const chosen = tSel ? [...tSel.selectedOptions].map(o => o.value).filter(Boolean) : [];

  const oSel    = document.getElementById("origin-select");
  const origins = oSel ? [...oSel.selectedOptions].map(o => o.value).filter(Boolean) : [];

  if (from) p.set("date_from", from);
  if (to)   p.set("date_to",   to);
  if (city) p.set("city",      city);
  if (chosen.length > 0 && chosen.length < allTypes.length)
    p.set("types", chosen.join(","));
  if (origins.length > 0 && origins.length < allOrigins.length)
    p.set("origins", origins.join(","));
  return p;
}

function applyFilters() {
  const p  = buildFilterParams();
  const qs = p.toString() ? "?" + p.toString() : "";
  document.getElementById("export-btn").href = "/export" + qs;
  updateActiveTag(p);
  detectPreset();
  refresh(p);
}

function clearFilters() {
  document.getElementById("f-from").value = "";
  document.getElementById("f-to").value   = "";
  document.getElementById("f-city").value = "";
  const tSel = document.getElementById("type-select");
  if (tSel) [...tSel.options].forEach(o => o.selected = false);
  const oSel = document.getElementById("origin-select");
  if (oSel) [...oSel.options].forEach(o => o.selected = false);
  applyFilters();
}

function selectAllTypes() {
  const sel = document.getElementById("type-select");
  if (sel) [...sel.options].forEach(o => o.selected = true);
  applyFilters();
}
function clearTypes() {
  const sel = document.getElementById("type-select");
  if (sel) [...sel.options].forEach(o => o.selected = false);
  applyFilters();
}
function selectAllOrigins() {
  const sel = document.getElementById("origin-select");
  if (sel) [...sel.options].forEach(o => o.selected = true);
  applyFilters();
}
function clearOrigins() {
  const sel = document.getElementById("origin-select");
  if (sel) [...sel.options].forEach(o => o.selected = false);
  applyFilters();
}

function updateActiveTag(p) {
  const parts = [];
  if (p.get("date_from") || p.get("date_to")) {
    const f = p.get("date_from") || ""; const t = p.get("date_to") || "";
    parts.push(f && t ? f.slice(5)+" — "+t.slice(5) : f||t);
  }
  if (p.get("city"))  parts.push(esc(p.get("city")));
  if (p.get("types"))   parts.push(p.get("types").split(",").length + " סוגים");
  if (p.get("origins")) parts.push(p.get("origins").split(",").join(" / "));
  const tag = document.getElementById("filter-active-tag");
  tag.textContent = parts.length ? "· " + parts.join(" · ") : "";
}

/* ── Load metadata ── */
async function loadCities() {
  const cities = await fetch("/api/cities").then(r => r.json());
  document.getElementById("city-list").innerHTML = cities.map(c => `<option value="${esc(c)}">`).join("");
}

async function loadTypes() {
  allTypes = await fetch("/api/types").then(r => r.json());
  const sel = document.getElementById("type-select");
  if (!allTypes.length) {
    sel.innerHTML = '<option value="" disabled>אין נתונים עדיין</option>';
    return;
  }
  sel.innerHTML = allTypes.map(t => {
    const cnt = t.total ? ` (${t.total.toLocaleString()})` : "";
    return `<option value="${esc(t.title)}">${esc(t.title)}${cnt}</option>`;
  }).join("");
}

async function loadOrigins() {
  allOrigins = await fetch("/api/origins").then(r => r.json());
  const sel = document.getElementById("origin-select");
  if (!allOrigins.length) {
    sel.innerHTML = '<option value="" disabled>אין נתונים</option>';
    return;
  }
  sel.innerHTML = allOrigins.map(o => {
    const label = ORIGIN_HE[o.origin] || o.origin;
    const cnt   = o.total ? ` (${o.total.toLocaleString()})` : "";
    return `<option value="${esc(o.origin)}">${esc(label)}${cnt}</option>`;
  }).join("");
}

/* ── Refresh ── */
async function refresh(filterParams) {
  const fp = filterParams || buildFilterParams();
  const qs = fp.toString() ? "?" + fp.toString() : "";
  const [stats, state, recent] = await Promise.all([
    fetch("/api/stats"  + qs).then(r => r.json()),
    fetch("/api/state").then(r => r.json()),
    fetch("/api/recent" + qs).then(r => r.json()),
  ]);

  document.getElementById("total").textContent     = stats.total.toLocaleString();
  document.getElementById("cities").textContent    = stats.cities.toLocaleString();
  document.getElementById("new_today").textContent = state.new_today.toLocaleString();
  document.getElementById("newest").textContent    = stats.newest ? stats.newest.slice(11,16) : "—";
  document.getElementById("last_live").textContent    = state.last_live;
  document.getElementById("last_history").textContent = state.last_history;
  document.getElementById("errors").textContent       = state.live_errors;

  // Result bar
  const rb = document.getElementById("result-bar");
  const isFiltered = qs.length > 0;
  if (stats.total === 0 && isFiltered) {
    rb.className = "result-bar empty";
    rb.textContent = "⚠️ אין תוצאות לפילטר הנוכחי";
  } else if (isFiltered) {
    rb.className = "result-bar filtered";
    rb.textContent = `📊 מציג ${stats.total.toLocaleString()} התראות (מסונן)`;
  } else {
    rb.className = "result-bar";
    rb.textContent = `📊 סה"כ ${stats.total.toLocaleString()} התראות`;
  }

  // Recent table
  const tbody = document.getElementById("recent");
  if (!recent.length) {
    tbody.innerHTML = `<tr class="empty-row"><td colspan="3">אין תוצאות לפילטר הנוכחי</td></tr>`;
  } else {
    tbody.innerHTML = "";
    recent.forEach(r => {
      const color = colorFor(r.title);
      tbody.innerHTML += `<tr>
        <td>${fmtDt(r.alert_dt)}</td>
        <td>${esc(r.city)}</td>
        <td><span class="badge" style="background:${color}">${esc(r.title)}</span></td>
      </tr>`;
    });
  }
}

/* ── Auto-apply with debounce on date/city changes ── */
function debouncedApply() {
  clearTimeout(debounceTimer);
  debounceTimer = setTimeout(applyFilters, 600);
}
document.getElementById("f-from").addEventListener("change", debouncedApply);
document.getElementById("f-to").addEventListener("change",   debouncedApply);
document.getElementById("f-city").addEventListener("input",  debouncedApply);
document.getElementById("type-select").addEventListener("change", applyFilters);
document.getElementById("origin-select").addEventListener("change", applyFilters);

/* ── Init ── */
let cd = 15;
setInterval(() => {
  cd--; document.getElementById("countdown").textContent = cd;
  if (cd <= 0) { cd = 15; refresh(); }
}, 1000);

loadCities();
loadOrigins();
loadTypes().then(() => {
  setPreset("all");   // מסמן "הכל" כברירת מחדל
});
</script>
</body>
</html>
"""

@app.route("/")
def index():
    return render_template_string(HTML, cat_colors=CAT_COLORS)

@app.route("/api/stats")
def api_stats():
    where, params = _parse_filters(request.args)
    s = get_stats(where, params)
    return jsonify(
        total=s["total"],
        cities=s["cities"],
        newest=s["newest"],
        oldest=s["oldest"],
    )

@app.route("/api/state")
def api_state():
    return jsonify(**_state)

@app.route("/api/recent")
def api_recent():
    where, params = _parse_filters(request.args)
    rows = query_alerts(where, params)[:10]
    return jsonify([dict(r) for r in rows])

@app.route("/api/cities")
def api_cities():
    with get_db() as conn:
        rows = conn.execute(
            "SELECT DISTINCT city FROM alerts ORDER BY city"
        ).fetchall()
    return jsonify([r["city"] for r in rows])

@app.route("/api/types")
def api_types():
    with get_db() as conn:
        rows = conn.execute(
            "SELECT title, MIN(category) as category, COUNT(*) as total "
            "FROM alerts GROUP BY title ORDER BY total DESC"
        ).fetchall()
    return jsonify([
        {"title": r["title"], "category": r["category"], "total": r["total"]}
        for r in rows
    ])

@app.route("/api/origins")
def api_origins():
    with get_db() as conn:
        rows = conn.execute(
            "SELECT origin, COUNT(*) as total FROM alerts "
            "WHERE origin != '' GROUP BY origin ORDER BY total DESC"
        ).fetchall()
    return jsonify([
        {"origin": r["origin"], "total": r["total"]} for r in rows
    ])

@app.route("/export")
def export():
    args  = request.args
    where, params = _parse_filters(args)

    # תיאור פילטר לגיליון הסטטיסטיקות
    parts = []
    if args.get("date_from"): parts.append(f"מתאריך {args['date_from']}")
    if args.get("date_to"):   parts.append(f"עד {args['date_to']}")
    if args.get("city"):      parts.append(f"יישוב: {args['city']}")
    if args.get("types"):
        types_list = [t.strip() for t in args["types"].split(",") if t.strip()]
        parts.append(f"סוגי התראות: {', '.join(types_list)}")
    if args.get("origins"):
        origins_list = [o.strip() for o in args["origins"].split(",") if o.strip()]
        parts.append(f"מקור: {', '.join(origins_list)}")
    filter_info = " | ".join(parts) if parts else None

    s     = get_stats(where, params)
    today = datetime.now().strftime("%d-%m-%Y")
    fname = f"התראות_פיקוד_העורף_{today}.xlsx"
    buf   = build_excel(where, params, filter_info)
    log.info("Excel exported: %s (%d alerts, filter=%s)", fname, s["total"], bool(filter_info))
    return send_file(
        buf,
        as_attachment=True,
        download_name=fname,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

def startup_backfill():
    """טוען בהפעלה:
    0. ארכיון CSV מ-GitHub (נתונים היסטוריים 2019–היום)
    1. היסטוריה סטטית — AlertsHistory.json (נתוני אתמול וקודם)
    2. GetAlarmsHistory — 3000 האחרונות (נתוני היום)
    """

    # 0. ארכיון GitHub CSV — מייבא רשומות חסרות בלבד (INSERT OR IGNORE)
    try:
        log.info("Backfill: מוריד ארכיון CSV מ-GitHub...")
        resp = requests.get(CSV_SOURCE_URL, timeout=60)
        resp.raise_for_status()
        text = resp.content.decode("utf-8", errors="replace")
        reader = csv.DictReader(io.StringIO(text))
        before = get_stats()["total"]
        import_csv_rows(reader)
        after  = get_stats()["total"]
        log.info("Backfill CSV: %d רשומות חדשות נוספו (סה\"כ %d)", after - before, after)
    except Exception as e:
        log.warning("Backfill CSV error: %s", e)

    # 1. היסטוריה סטטית — AlertsHistory.json (נתוני אתמול וקודם)
    try:
        log.info("Backfill: טוען AlertsHistory.json...")
        data = fetch_json(STATIC_HISTORY_URL, headers=OREF_HEADERS)
        n = insert_alerts(data, source="history")
        log.info("Backfill: נטענו %d רשומות מ-AlertsHistory.json", n)
    except Exception as e:
        log.warning("Backfill AlertsHistory error: %s", e)

    # 2. GetAlarmsHistory — 3000 האחרונות (נתוני היום)
    try:
        log.info("Backfill: טוען GetAlarmsHistory.aspx (נתוני היום)...")
        data = fetch_json(HISTORY_URL, headers=HISTORY_HEADERS)
        n = insert_alerts(data, source="history")
        log.info("Backfill: נטענו %d רשומות מ-GetAlarmsHistory.aspx", n)
    except Exception as e:
        log.warning("Backfill GetAlarmsHistory error: %s", e)

    _state["last_history"] = datetime.now().strftime("%H:%M:%S")
    total = get_stats()["total"]
    log.info("Backfill הושלם — סה\"כ %d רשומות ב-DB", total)


# ─── Main ─────────────────────────────────────────────────────────────────────
def main():
    import os
    global DB_PATH, IS_CLOUD

    # בענן: PORT מגיע מ-environment variable; מקומית: 5050
    port     = int(os.environ.get("PORT", PORT))
    IS_CLOUD = bool(os.environ.get("RENDER") or os.environ.get("RAILWAY_ENVIRONMENT"))

    script_dir = os.path.dirname(os.path.abspath(__file__))
    db_dir     = os.environ.get("DB_DIR", script_dir)   # בענן אפשר לדרוס נתיב
    DB_PATH = os.path.join(db_dir, "alerts.db")

    os.makedirs(db_dir, exist_ok=True)
    init_db()

    # טעינה ראשונית
    threading.Thread(target=startup_backfill, name="backfill", daemon=True).start()

    # Collector היסטוריה — פועל תמיד (בענן ומקומית)
    threading.Thread(target=collect_history, name="history", daemon=True).start()

    # Collector Live (Alerts.json) — רק מקומית; בענן oref.org.il חוסם IP של שרתי ענן
    if not IS_CLOUD:
        threading.Thread(target=collect_live, name="live", daemon=True).start()
        log.info("Live collector enabled (local mode)")
    else:
        log.info("Live collector disabled (cloud mode — GetAlarmsHistory covers recent data)")
        _state["last_live"] = "N/A (ענן)"

    log.info("Server starting on port %d (cloud=%s)", port, IS_CLOUD)

    # פתיחת דפדפן רק במצב מקומי
    if not IS_CLOUD:
        import webbrowser
        threading.Timer(1.5, lambda: webbrowser.open(f"http://localhost:{port}")).start()

    app.run(host="0.0.0.0", port=port, debug=False, use_reloader=False)


if __name__ == "__main__":
    main()
